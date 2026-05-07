"""
核销工具 API 路由 (集成到 o2-workbench)
"""
import os, json, re
from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel
import openpyxl

router = APIRouter(prefix='/api/hexiao', tags=['核销'])

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
REMOTE = ['新疆', '西藏', '青海', '内蒙', '内蒙古']
PRICES_FILE = os.path.join(BASE, 'hexiao', 'prices.json')

def load_prices():
    try:
        if os.path.exists(PRICES_FILE):
            with open(PRICES_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except: pass
    return {'prices': [], 'shipping': {}}
def save_prices(data):
    with open(PRICES_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def match(shop, sn, prices, shipping):
    s = shipping.get(shop, {'n':2.5,'r':10})
    for p in prices:
        if p['shop']==shop and (p['k1']==sn or sn.startswith(p['k1']) or p['k1'].startswith(sn)):
            # Use per-entry shipping if available, otherwise customer default
            ns = p.get('n', s['n'])
            rs = p.get('r', s['r'])
            return p['price'], p['code'], ns, rs
    return None, None, s['n'], s['r']

def process(delivery_path):
    pd = load_prices()
    wb = openpyxl.load_workbook(delivery_path, data_only=True)
    ws = wb.active
    rows, ls, lf, la = [], '', '', ''
    for r in range(2, ws.max_row + 1):
        a = str(ws.cell(r,5).value or '').strip()
        s = str(ws.cell(r,10).value or '').strip()
        f = str(ws.cell(r,11).value or '').strip()
        n = str(ws.cell(r,23).value or '').strip()
        sp = str(ws.cell(r,24).value or '').strip()
        if not s and not n and not sp: continue
        if s: ls, lf, la = s, f, a
        rows.append({'shop':ls,'fac':lf,'name':n,'spec':sp,'addr':la})
    pl, sh = pd.get('prices',[]), pd.get('shipping',{})
    res = []
    for row in rows:
        m = re.match(r'(\d+)', row['spec'] or ''); q = int(m.group(1)) if m else 1
        rm = any(k in row['addr'] for k in REMOTE)
        pr, cd, nr, rr = match(row['shop'], row['name'], pl, sh)
        sc = rr if rm else nr; gd = round(q*pr,2) if pr else 0
        res.append({'shop':row['shop'],'fac':row['fac'],'name':row['name'],'spec':row['spec'],
            'qty':q,'price':pr,'code':cd,'goods':gd,'remote':rm,'ship':sc,'total':round(gd+sc,2),'match':pr is not None})
    res.sort(key=lambda x: x['shop'])
    stats = {'total': len(res), 'matched': sum(1 for r in res if r['price']),
        'goods_total': round(sum(r['goods'] for r in res),2),
        'ship_total': round(sum(r['ship'] for r in res),2),
        'grand_total': round(sum(r['total'] for r in res),2)}
    return {'results': res, 'stats': stats}

# ---- API endpoints ----
@router.get('/prices')
def get_prices():
    return load_prices()

@router.post('/prices/import')
def import_prices():
    paths = [os.path.join(BASE, '对账单5.4.xlsx'), os.path.join(BASE, 'hexiao', '对账单5.4.xlsx')]
    found = None
    for p in paths:
        if os.path.exists(p): found = p; break
    if not found:
        raise HTTPException(404, '未找到对账单模板文件')
    wb = openpyxl.load_workbook(found, data_only=True); ws = wb['汇总']
    prices, shipping, cur = [], {}, None
    for r in range(3, ws.max_row + 1):
        cust = ws.cell(r, 3).value
        if cust:
            cur = str(cust).strip()
            if cur not in shipping:
                shipping[cur] = {'n': float(ws.cell(r,12).value or 2.5), 'r': float(ws.cell(r,14).value or 10)}
        if cur and ws.cell(r, 4).value:
            k1 = str(ws.cell(r,4).value).strip()
            ns = float(ws.cell(r,12).value) if ws.cell(r,12).value else (shipping.get(cur,{}).get('n',2.5))
            rs = float(ws.cell(r,14).value) if ws.cell(r,14).value else (shipping.get(cur,{}).get('r',10))
            if not any(x['shop']==cur and x['k1']==k1 for x in prices):
                prices.append({'shop':cur,'k1':k1,'code':str(ws.cell(r,6).value or '').strip(),
                    'price':float(ws.cell(r,9).value) if ws.cell(r,9).value else None,
                    'n': ns, 'r': rs})
    data = {'prices': prices, 'shipping': shipping}
    save_prices(data)
    return data

@router.post('/prices/save')
def save_prices_api(data: dict):
    save_prices(data)
    return {'ok': True, 'count': len(data.get('prices',[]))}

@router.post('/process')
async def process_file(file: UploadFile = File(...)):
    tmp = os.path.join(BASE, 'hexiao', '_upload.xlsx')
    content = await file.read()
    with open(tmp, 'wb') as f: f.write(content)
    result = process(tmp)
    os.remove(tmp)
    return result
