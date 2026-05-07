"""
核销工具 — 提取·对应·总结
"""
import sys, os, json, re
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTextEdit, QProgressBar, QPushButton, QLabel, QFileDialog,
    QMessageBox, QLineEdit, QTableWidget, QTableWidgetItem,
    QHeaderView, QFrame, QAbstractItemView, QDialog, QComboBox,
    QSizePolicy, QGraphicsDropShadowEffect
)
from PySide6.QtCore import Signal, Qt, QThread
from PySide6.QtGui import QFont, QColor
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

APP_DIR = os.path.dirname(os.path.abspath(__file__))
PRICES_FILE = os.path.join(APP_DIR, 'prices.json')

# ===== data =====
def load_prices():
    try:
        if os.path.exists(PRICES_FILE):
            with open(PRICES_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except: pass
    return {'prices': [], 'shipping': {}}
def save_prices(data):
    with open(PRICES_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
def import_from_template(path):
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb['汇总']
    prices, shipping, cur = [], {}, None
    for r in range(3, ws.max_row + 1):
        cust = ws.cell(r, 3).value
        if cust:
            cur = str(cust).strip()
            if cur not in shipping:
                shipping[cur] = {'n': float(ws.cell(r,12).value or 2.5), 'r': float(ws.cell(r,14).value or 10)}
        if cur and ws.cell(r, 4).value:
            k1 = str(ws.cell(r,4).value).strip()
            if not any(x['shop']==cur and x['k1']==k1 for x in prices):
                prices.append({'shop':cur,'k1':k1,'code':str(ws.cell(r,6).value or '').strip(),
                    'price':float(ws.cell(r,9).value) if ws.cell(r,9).value else None})
    return {'prices':prices, 'shipping':shipping}
def match(shop, sn, prices, shipping):
    s = shipping.get(shop, {'n':2.5,'r':10})
    for p in prices:
        if p['shop']==shop and (p['k1']==sn or sn.startswith(p['k1']) or p['k1'].startswith(sn)):
            return p['price'], p['code'], s['n'], s['r']
    return None, None, s['n'], s['r']

# ===== core =====
REMOTE = ['新疆','西藏','青海','内蒙','内蒙古']
def extract(path):
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb.active
    rows, ls, lf, la = [], '', '', ''
    for r in range(2, ws.max_row + 1):
        a = str(ws.cell(r,5).value or '').strip(); s = str(ws.cell(r,10).value or '').strip()
        f = str(ws.cell(r,11).value or '').strip(); n = str(ws.cell(r,23).value or '').strip()
        sp = str(ws.cell(r,24).value or '').strip()
        if not s and not n and not sp: continue
        if s: ls, lf, la = s, f, a
        rows.append({'shop':ls,'fac':lf,'name':n,'spec':sp,'addr':la})
    return rows
def qty(s):
    m = re.match(r'(\d+)', s or ''); return int(m.group(1)) if m else 1
def is_remote(a):
    return any(k in a for k in REMOTE)

def run(delivery, pd):
    rows = extract(delivery); pl, sh = pd.get('prices',[]), pd.get('shipping',{}); res = []
    for row in rows:
        q = qty(row['spec']); rm = is_remote(row['addr'])
        pr, cd, nr, rr = match(row['shop'], row['name'], pl, sh)
        sc = rr if rm else nr; gd = round(q*pr,2) if pr else 0
        res.append({'shop':row['shop'],'fac':row['fac'],'name':row['name'],'spec':row['spec'],
            'qty':q,'price':pr,'code':cd or '','goods':gd,'remote':rm,'ship':sc,'total':round(gd+sc,2)})
    res.sort(key=lambda x: x['shop']); return res

def out_excel(res, path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '核销汇总'
    ws.cell(1,1,'对账单核销汇总').font = Font(size=14,bold=True); ws.merge_cells('A1:K1')
    hd = ['商家/店铺','厂家','商品简称','规格','数量','单价','货款','偏远','运费','合计','匹配编码']
    hf = PatternFill(start_color='B4C6E7',end_color='B4C6E7',fill_type='solid')
    for i,h in enumerate(hd,1):
        c=ws.cell(2,i,h); c.fill=hf; c.font=Font(bold=True)
    for ri,r in enumerate(res,3):
        for ci,v in [(1,r['shop']),(2,r['fac']),(3,r['name']),(4,r['spec']),(5,r['qty']),
                     (6,r['price'] if r['price'] else ''),(7,r['goods']),
                     (8,'偏远' if r['remote'] else ''),(9,r['ship']),(10,r['total']),(11,r['code'])]:
            ws.cell(ri,ci,v)
    for i,w in enumerate([30,30,22,16,6,8,10,6,8,10,22],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    tb = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    for rw in range(2,len(res)+3):
        for cl in range(1,12): ws.cell(rw,cl).border=tb
    wb.save(path)

# ===== worker =====
class Worker(QThread):
    log_signal = Signal(str)
    done_signal = Signal(bool,str)
    def __init__(self, d, pd, o):
        super().__init__(); self.d=d; self.pd=pd; self.o=o
    def run(self):
        try:
            res = run(self.d, self.pd); m = sum(1 for r in res if r['price'])
            self.log_signal.emit(f'{len(res)} 单，匹配 {m} 单单价')
            self.log_signal.emit(f'货款 {round(sum(r["goods"] for r in res),2)} · 运费 {round(sum(r["ship"] for r in res),2)} · 合计 {round(sum(r["total"] for r in res),2)}')
            out_excel(res, self.o)
            self.log_signal.emit(f'已保存: {self.o}')
            self.done_signal.emit(True, f'完成！{len(res)} 条')
        except Exception as e:
            import traceback; self.log_signal.emit(str(e)); self.log_signal.emit(traceback.format_exc())
            self.done_signal.emit(False, str(e))

# ===== price dialog =====
class PriceDialog(QDialog):
    def __init__(self, data, parent=None):
        super().__init__(parent)
        self.setWindowTitle('编辑价格'); self.resize(720, 500)
        self.data = data; self._ui()
    def _ui(self):
        self.setStyleSheet('QDialog{background:#f8fafc;}')
        l = QVBoxLayout(self); l.setContentsMargins(20,20,20,20)
        bar = QHBoxLayout()
        flt_label = QLabel('按客户筛选:'); flt_label.setStyleSheet('font-size:13px;')
        bar.addWidget(flt_label)
        self.fc = QComboBox(); self.fc.setStyleSheet('max-width:200px; padding:4px 8px;')
        self.fc.currentIndexChanged.connect(self._refill); bar.addWidget(self.fc)
        bar.addStretch()
        for txt, cb in [('+ 新增', self._add), ('— 删除', self._del)]:
            b = QPushButton(txt); b.setStyleSheet('padding:6px 16px;'); b.clicked.connect(cb); bar.addWidget(b)
        imp = QPushButton('从模板导入'); imp.setStyleSheet('padding:6px 16px;'); imp.clicked.connect(self._import); bar.addWidget(imp)
        l.addLayout(bar)
        self.t = QTableWidget()
        self.t.setColumnCount(6)
        self.t.setHorizontalHeaderLabels(['客户', '商品简称', '匹配编码', '单价', '普通运费', '偏远运费'])
        for i in [0,1]: self.t.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)
        for i in [2,3,4,5]: self.t.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeToContents)
        self.t.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.t.setAlternatingRowColors(True)
        l.addWidget(self.t, 1)
        h = QHBoxLayout(); h.addStretch()
        sv = QPushButton('保存'); sv.setStyleSheet('padding:8px 32px; background:#3b82f6; color:white; border:none; border-radius:6px; font-weight:600;')
        sv.clicked.connect(self._save); h.addWidget(sv); l.addLayout(h); self._refill()
    def _refill(self):
        pl = self.data.get('prices',[]); sh = self.data.get('shipping',{})
        shops = sorted(set(p['shop'] for p in pl))
        self.fc.blockSignals(True); cur = self.fc.currentText()
        self.fc.clear(); self.fc.addItem('全部客户'); self.fc.addItems(shops)
        idx = self.fc.findText(cur)
        if idx>=0: self.fc.setCurrentIndex(idx)
        self.fc.blockSignals(False)
        flt = self.fc.currentText()
        filtered = [p for p in pl if flt=='全部客户' or p['shop']==flt]
        self.t.setRowCount(len(filtered))
        for i,p in enumerate(filtered):
            s = sh.get(p['shop'],{'n':2.5,'r':10})
            for j,v in [(0,p['shop']),(1,p['k1']),(2,p.get('code','')),(3,str(p['price']) if p['price'] else ''),(4,str(s['n'])),(5,str(s['r']))]:
                self.t.setItem(i,j,QTableWidgetItem(v))
    def _add(self):
        self.data.setdefault('prices',[]).append({'shop':'','k1':'','code':'','price':None}); self._refill()
    def _del(self):
        rows = sorted(set(i.row() for i in self.t.selectedIndexes()), reverse=True)
        pl = self.data.get('prices',[])
        for r in rows:
            if r < len(pl): pl.pop(r)
        self._refill()
    def _import(self):
        p,_ = QFileDialog.getOpenFileName(self,'选择对账单模板','','Excel (*.xlsx)')
        if p: self.data = import_from_template(p); self._refill(); QMessageBox.information(self,'导入完成',f'{len(self.data["prices"])} 条')
    def _save(self):
        pl = self.data.setdefault('prices',[]); sh = self.data.setdefault('shipping',{})
        for i in range(self.t.rowCount()):
            shop = self.t.item(i,0).text().strip() if self.t.item(i,0) else ''
            k1 = self.t.item(i,1).text().strip() if self.t.item(i,1) else ''
            code = self.t.item(i,2).text().strip() if self.t.item(i,2) else ''
            ps = self.t.item(i,3).text().strip() if self.t.item(i,3) else ''
            ns = self.t.item(i,4).text().strip() if self.t.item(i,4) else ''
            rs = self.t.item(i,5).text().strip() if self.t.item(i,5) else ''
            p = float(ps) if ps else None; n = float(ns) if ns else 2.5; r = float(rs) if rs else 10
            if i < len(pl): pl[i] = {'shop':shop,'k1':k1,'code':code,'price':p}
            if shop: sh[shop] = {'n':n,'r':r}
        save_prices(self.data); QMessageBox.information(self,'已保存',f'{len(pl)} 条')

# ===== main window =====
class MainWin(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('核销工具')
        self.resize(960, 640)
        self.pd = load_prices()
        self._ui()

    def _card(self, w, extra=''):
        c = QFrame(); c.setStyleSheet(f'QFrame{{background:white; border-radius:12px; border:1px solid #e8ecf0;}}{extra}')
        c.setGraphicsEffect(QGraphicsDropShadowEffect(blurRadius=20, offset=QPoint(0,2), color=QColor(0,0,0,15)))
        QVBoxLayout(c).setContentsMargins(20,20,20,20) if w else None
        return c

    def _ui(self):
        central = QWidget(); self.setCentralWidget(central)
        root = QVBoxLayout(central); root.setContentsMargins(0,0,0,0); root.setSpacing(0)

        # title
        tb = QFrame(); tb.setStyleSheet('background:white; border-bottom:1px solid #e8ecf0;')
        tbh = QHBoxLayout(tb); tbh.setContentsMargins(28,14,28,14)
        ti = QVBoxLayout(); ti.setSpacing(0)
        t1 = QLabel('核销工具'); t1.setStyleSheet('font-size:17px; font-weight:700; color:#1e293b;')
        t2 = QLabel('从发货明细自动生成对账单'); t2.setStyleSheet('font-size:11px; color:#94a3b8;')
        ti.addWidget(t1); ti.addWidget(t2); tbh.addLayout(ti); tbh.addStretch()
        vb = QLabel('v3'); vb.setStyleSheet('background:#eff6ff; color:#3b82f6; padding:2px 14px; border-radius:8px; font-size:11px; font-weight:600;')
        tbh.addWidget(vb); root.addWidget(tb)

        # body
        body = QWidget(); body.setStyleSheet('background:#f1f5f9;')
        bl = QHBoxLayout(body); bl.setContentsMargins(28,24,28,24); bl.setSpacing(24)

        # left panel
        left = QFrame(); left.setStyleSheet('QFrame{background:white; border-radius:12px; border:1px solid #e8ecf0;}')
        left.setFixedWidth(280)
        ll = QVBoxLayout(left); ll.setContentsMargins(20,20,20,20)

        ll.addWidget(QLabel('价格数据'))
        # stats
        stats = QFrame(); stats.setStyleSheet('background:#f8fafc; border-radius:10px;')
        stl = QHBoxLayout(stats); stl.setContentsMargins(16,12,16,12)
        self.stat_left = QLabel('0'); self.stat_left.setStyleSheet('font-size:22px; font-weight:700; color:#1e293b;')
        self.stat_right = QLabel('0'); self.stat_right.setStyleSheet('font-size:22px; font-weight:700; color:#1e293b;')
        sl = QVBoxLayout(); sl.setSpacing(0); sl.addWidget(self.stat_left); sl.addWidget(QLabel('条价格')); sl.setSpacing(0); stl.addLayout(sl)
        stl.addStretch()
        sr = QVBoxLayout(); sr.setSpacing(0); sr.addWidget(self.stat_right); sr.addWidget(QLabel('家运费')); stl.addLayout(sr)
        ll.addWidget(stats)

        pe = QPushButton('编辑价格')
        pe.setStyleSheet('QPushButton{background:#3b82f6; color:white; border:none; border-radius:8px; padding:10px 0; font-weight:600;} QPushButton:hover{background:#2563eb;}')
        pe.clicked.connect(self._edit); ll.addWidget(pe)

        self.pv = QTextEdit(); self.pv.setReadOnly(True)
        self.pv.setStyleSheet('QTextEdit{border:1px solid #e2e8f0; border-radius:8px; background:#fafbfc; font-size:11px; padding:8px;}')
        ll.addWidget(self.pv, 1); bl.addWidget(left)

        # right panel
        right = QFrame(); right.setStyleSheet('QFrame{background:white; border-radius:12px; border:1px solid #e8ecf0;}')
        rl = QVBoxLayout(right); rl.setContentsMargins(24,24,24,24)
        rl.addWidget(QLabel('处理'))

        # file upload drop zone
        self.drop_frame = QFrame()
        self.drop_frame.setStyleSheet('QFrame{background:#f8fafc; border:2px dashed #cbd5e1; border-radius:12px;} QFrame:hover{border-color:#3b82f6; background:#eff6ff;}')
        self.drop_frame.setAcceptDrops(False)
        self.drop_frame.setFixedHeight(90)
        dl = QHBoxLayout(self.drop_frame); dl.setContentsMargins(16,0,16,0)
        ic = QLabel('📁'); ic.setStyleSheet('font-size:28px;'); dl.addWidget(ic)
        self.file_label = QLabel('点击选择或拖入发货明细 Excel')
        self.file_label.setStyleSheet('font-size:13px; color:#64748b;'); dl.addWidget(self.file_label, 1)
        self.file_btn = QPushButton('选择文件'); self.file_btn.setStyleSheet('padding:8px 20px; border:1px solid #e2e8f0; border-radius:8px;')
        self.file_btn.clicked.connect(self._sel_de); dl.addWidget(self.file_btn)
        rl.addWidget(self.drop_frame)

        # output path
        oh = QHBoxLayout()
        ol = QLabel('输出'); ol.setStyleSheet('min-width:36px; color:#64748b;')
        oh.addWidget(ol)
        self.oe = QLineEdit(); self.oe.setPlaceholderText('默认：发货明细同目录')
        self.oe.setStyleSheet('QLineEdit{padding:8px 12px; border:1px solid #e2e8f0; border-radius:8px; background:#fafbfc;}')
        oh.addWidget(self.oe)
        ob = QPushButton('选择目录'); ob.setStyleSheet('padding:8px 16px; border:1px solid #e2e8f0; border-radius:8px;')
        ob.clicked.connect(self._sel_oe); oh.addWidget(ob)
        rl.addLayout(oh)

        # progress
        self.sl = QLabel('就绪'); self.sl.setStyleSheet('color:#94a3b8; font-size:12px;'); rl.addWidget(self.sl)
        self.pb = QProgressBar(); self.pb.setFixedHeight(8)
        self.pb.setStyleSheet('QProgressBar{border:none; border-radius:4px; background:#e2e8f0; height:8px; color:transparent;} QProgressBar::chunk{background:#3b82f6; border-radius:4px;}')
        rl.addWidget(self.pb)

        self.log = QTextEdit(); self.log.setReadOnly(True)
        self.log.setStyleSheet('QTextEdit{background:#1e293b; color:#a5f3fc; border:none; border-radius:8px; padding:12px; font-size:12px;}')
        rl.addWidget(self.log, 1); bl.addWidget(right, 1); root.addWidget(body, 1)

        # status bar
        sb = QFrame(); sb.setStyleSheet('background:white; border-top:1px solid #e8ecf0;')
        sbl = QHBoxLayout(sb); sbl.setContentsMargins(28,10,28,10)
        self.st = QLabel('待处理'); self.st.setStyleSheet('color:#94a3b8; font-size:12px;'); sbl.addWidget(self.st)
        sbl.addStretch()
        self.sb = QPushButton('开始处理')
        self.sb.setStyleSheet('QPushButton{background:#3b82f6; color:white; border:none; border-radius:8px; padding:8px 28px; font-weight:600;} QPushButton:hover{background:#2563eb;} QPushButton:disabled{background:#94a3b8;}')
        self.sb.clicked.connect(self._start); sbl.addWidget(self.sb); root.addWidget(sb)
        self._refresh()

    def _sel_de(self):
        p,_ = QFileDialog.getOpenFileName(self,'选择发货明细','','Excel (*.xlsx *.xls)')
        if p:
            self.de = p
            self.file_label.setText(os.path.basename(p))
            self.file_label.setStyleSheet('font-size:13px; color:#1e293b; font-weight:500;')
            self.drop_frame.setStyleSheet('QFrame{background:#eff6ff; border:2px solid #3b82f6; border-radius:12px;}')
            self.oe.setText(os.path.join(os.path.dirname(p),'对账单汇总表.xlsx'))
    def _sel_oe(self):
        p = QFileDialog.getExistingDirectory(self,'选择输出目录')
        if p: self.oe.setText(os.path.join(p,'对账单汇总表.xlsx'))
    def _edit(self):
        dlg = PriceDialog(self.pd, self)
        if dlg.exec(): self.pd = dlg.data; self._refresh()
    def _refresh(self):
        pl = self.pd.get('prices',[]); sh = self.pd.get('shipping',{})
        self.stat_left.setText(str(len(pl))); self.stat_right.setText(str(len(sh)))
        self.st.setText(f'价格: {len(pl)} 条 | 运费: {len(sh)} 家')
        lines = [f'{p["shop"][:18]:20s} {p["k1"][:12]:15s} ¥{p["price"]}' for p in pl[:18]]
        if len(pl) > 18: lines.append(f'... 还有 {len(pl)-18} 条')
        self.pv.setText('\n'.join(lines))
    def _start(self):
        if not hasattr(self,'de') or not self.de or not os.path.exists(self.de):
            QMessageBox.warning(self,'提示','请先选择发货明细文件'); return
        o = self.oe.text().strip() or os.path.join(os.path.dirname(self.de),'对账单汇总表.xlsx')
        self.sb.setEnabled(False); self.pb.setValue(0); self.log.clear()
        self.log.append(f'📄 {self.de}')
        self.log.append(f'📊 {len(self.pd.get("prices",[]))} 条价格 · {len(self.pd.get("shipping",{}))} 家运费')
        w = Worker(self.de, self.pd, o)
        w.log_signal.connect(lambda m: (self.log.append(m),
            self.log.verticalScrollBar().setValue(self.log.verticalScrollBar().maximum())))
        w.done_signal.connect(lambda ok,msg: (
            setattr(self.sb,'enabled',True), setattr(self.pb,'value',100),
            self.sl.setText('完成' if ok else '失败'),
            QMessageBox.information(self,'完成',msg) if ok else QMessageBox.critical(self,'错误',msg)))
        w.start()

if __name__ == '__main__':
    QApplication.setStyle('Fusion')
    app = QApplication(sys.argv)
    app.setFont(QFont('Microsoft YaHei UI', 9))
    w = MainWin(); w.show(); sys.exit(app.exec())
